import os
from datetime import datetime
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from sqlalchemy import select

from app.models.request import Request, RequestItem
from app.models.user import User as Buyer
from app.models.counterparty import Counterparty


class ExcelProcessor:
    def __init__(self):
        # Определяем базовую директорию, где запущен скрипт
        base_dir = os.getcwd()
        self.outgoing_dir = os.path.join(base_dir, "excel_outgoing")
        self.incoming_dir = os.path.join(base_dir, "excel_incoming")
        
        # Создаем директории, если они не существуют
        os.makedirs(self.outgoing_dir, exist_ok=True)
        os.makedirs(self.incoming_dir, exist_ok=True)
        
    async def generate_request_excel(self, request_id: int, db: AsyncSession) -> Dict[str, str]:
        query = (
            select(Request)
            .where(Request.id == request_id)
            .options(
                selectinload(Request.items), 
                selectinload(Request.counterparty)
            )
        )
        result = await db.execute(query)
        request = result.scalar_one_or_none()
        
        if not request:
            raise ValueError(f"Request {request_id} not found")
        
        buyer_query = select(Buyer).where(Buyer.id == request.buyer_id)
        buyer_result = await db.execute(buyer_query)
        buyer = buyer_result.scalar_one_or_none()
        
        if not buyer:
            raise ValueError(f"Buyer {request.buyer_id} not found")
        
        # Группируем товары по kind ('metal' или 'generic')
        items_by_kind = {}
        for item in request.items:
            kind = item.kind or "generic" # Если kind не указан, считаем его generic
            if kind not in items_by_kind:
                items_by_kind[kind] = []
            items_by_kind[kind].append(item)
        
        files_created = {}
        
        for kind, items in items_by_kind.items():
            # Имя файла теперь основано на 'kind'
            filename = f"request_{request_id}_{kind}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.outgoing_dir, filename)
            
            self._create_category_excel(filepath, request, buyer, kind, items)
            
            files_created[kind] = filepath
        
        return files_created

    def _create_category_excel(self, filepath: str, request: Request, buyer: Buyer, kind: str, items: List[RequestItem]):
        wb = Workbook()
        ws = wb.active
        ws.title = kind[:31]

        ws.append(["Название поставки:", request.comment or ""])
        ws.append(["Дата и время поставки:", request.delivery_at.strftime("%Y-%m-%d %H:%M") if request.delivery_at else ""])
        ws.append(["Адрес поставки:", request.delivery_address or ""])
        counterparty_info = "Не указан"
        if request.counterparty:
            counterparty_info = f"{request.counterparty.short_name} (ИНН: {request.counterparty.inn})"
        ws.append(["Контрагент:", counterparty_info])
        ws.append([])

        # Теперь определяем тип таблицы по 'kind', а не по имени категории
        is_metal = (kind == "metal")
        # Название листа в Excel теперь будет 'Металлопрокат' или 'Прочее'
        ws.title = "Металлопрокат" if is_metal else "Прочее"

        if is_metal:
            headers = ["№", "Категория", "Размер", "ГОСТ", "Марка", "Аналоги", "Количество", "Комментарий", "Цена"]
        else:
            headers = ["№", "Наименование товаров/работ/услуг", "Размеры, характеристики", "Ед. изм.", "Количество", "Комментарий", "Цена"]
        
        ws.append(headers)

        for idx, item in enumerate(items, 1):
            if is_metal:
                row_data = [
                    idx,
                    item.category or "",
                    item.size or "",
                    item.state_standard or "",
                    item.stamp or "",
                    "Да" if item.allow_analogs else "Нет",
                    item.quantity,
                    item.comment or "",
                    ""
                ]
            else: # Generic
                row_data = [
                    idx,
                    item.name or "",
                    item.dims or "",
                    item.uom or "",
                    item.quantity,
                    item.comment or "",
                    ""
                ]
            ws.append(row_data)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[6]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(filepath)
    
    def read_pricing_excel(self, filepath: str) -> List[Dict[str, Any]]:
        try:
            # Читаем файл, пропуская первые 5 строк с информацией о заявке.
            # 6-я строка (индекс 5) будет использоваться как заголовок.
            df = pd.read_excel(filepath, header=5)
            pricing_data = []
            
            for index, row in df.iterrows():
                price_data = {
                    "row_number": index + 7, # +5 пропущенных строк +1 заголовок +1 для 1-based индексации
                    "price": None,
                    "raw_data": row.to_dict()
                }
                
                for col in df.columns:
                    if "цена" in str(col).lower() or "price" in str(col).lower():
                        try:
                            price_value = row[col]
                            # Проверяем, что значение не является NaN и не пустое
                            if pd.notna(price_value) and price_value != "":
                                price_data["price"] = float(price_value)
                        except (ValueError, TypeError):
                            # Если не удалось преобразовать в float, оставляем price=None
                            pass 
                
                pricing_data.append(price_data)
            
            return pricing_data
            
        except Exception as e:
            raise ValueError(f"Ошибка чтения Excel файла: {str(e)}")
    
    def get_available_files(self, directory: str) -> List[str]:
        if directory == "outgoing":
            target_dir = self.outgoing_dir
        elif directory == "incoming":
            target_dir = self.incoming_dir
        else:
            raise ValueError("Directory must be 'outgoing' or 'incoming'")
        
        if not os.path.exists(target_dir):
            return []
        
        files = []
        for filename in os.listdir(target_dir):
            if filename.endswith(('.xlsx', '.xls')):
                files.append(filename)
        
        return sorted(files)
